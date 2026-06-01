"""
excel_extractor.py — Extracción de datos financieros desde Excel (.xlsx / .xls)
Soporta:
  - Exportación PREVIRED (nómina cotizaciones)
  - Cartolas bancarias en Excel (BancoEstado, Banco de Chile)
  - Planillas de honorarios enviadas por contratistas
  - Nóminas SERVEL en formato Excel

Confianza esperada: 0.85–0.95 (datos estructurados en columnas)
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from typing import Optional, Any

import openpyxl
from io import BytesIO

from config import TipoDocumento, RUT_PARTIDO

logger = logging.getLogger(__name__)

RE_RUT   = re.compile(r"\b(\d{1,2}\.?\d{3}\.?\d{3}-[\dkK])\b")
RE_MONTO = re.compile(r"^[\$\s]*([\d.,]+)\s*$")


def parsear_excel(contenido: bytes, nombre_archivo: str = "") -> list[dict]:
    """
    Parsea un archivo Excel con datos financieros.

    Returns:
        Lista de dicts, uno por fila de dato relevante encontrada.
        Cada dict tiene los campos estándar + confidence_score.
    """
    resultado_base = _resultado_vacio(nombre_archivo)

    try:
        wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Error abriendo Excel %s: %s", nombre_archivo, e)
        resultado_base["confidence_score"] = 0.0
        resultado_base["error"] = str(e)
        return [resultado_base]

    # Detectar tipo por nombre de archivo o contenido de primera hoja
    tipo = _detectar_tipo(nombre_archivo, wb)
    resultado_base["tipo"] = tipo

    if tipo == TipoDocumento.PREVIRED:
        return _parsear_previred(wb, nombre_archivo)
    elif tipo == TipoDocumento.CARTOLA:
        return _parsear_cartola(wb, nombre_archivo)
    else:
        return _parsear_generico(wb, resultado_base, nombre_archivo)


# ─────────────────────────────────────────────────────────────────────
# PREVIRED — Exportación de cotizaciones previsionales
# ─────────────────────────────────────────────────────────────────────

def _parsear_previred(wb: openpyxl.Workbook, nombre_archivo: str) -> list[dict]:
    """
    Parsea exportación PREVIRED.
    Columnas esperadas: RUT | Nombre | AFP | Isapre | Bruto | Cotización total | Período
    """
    resultados = []
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [_resultado_vacio(nombre_archivo)]

    # Detectar fila de encabezados (primera fila con "RUT" o "NOMBRE")
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "rut" in row_str and ("nombre" in row_str or "trabajador" in row_str):
            header_idx = i
            break

    headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

    # Mapear columnas
    col_rut    = _find_col(headers, ["rut", "rut trabajador"])
    col_nombre = _find_col(headers, ["nombre", "nombre trabajador", "apellidos y nombre"])
    col_bruto  = _find_col(headers, ["sueldo bruto", "renta bruta", "base imponible", "bruto"])
    col_total  = _find_col(headers, ["total cotizaciones", "total", "monto total"])
    col_periodo = _find_col(headers, ["periodo", "período", "mes"])

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue

        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        rut_raw = str(cell(col_rut) or "").strip()
        if not rut_raw or rut_raw == "None":
            continue

        rut = _normalizar_rut(rut_raw)
        nombre = str(cell(col_nombre) or "").strip()
        bruto = _safe_int(cell(col_bruto))
        total_cot = _safe_int(cell(col_total))
        periodo_raw = cell(col_periodo)

        r = _resultado_vacio(nombre_archivo)
        r["tipo"]            = TipoDocumento.PREVIRED
        r["rut_emisor"]      = rut
        r["nombre_emisor"]   = nombre
        r["rut_receptor"]    = _normalizar_rut(RUT_PARTIDO)
        r["monto_bruto"]     = bruto
        r["monto_retencion"] = total_cot
        r["monto_liquido"]   = (bruto - total_cot) if (bruto and total_cot) else None
        r["concepto"]        = f"Cotizaciones previsionales {nombre}"
        r["fecha_documento"] = _parse_periodo(periodo_raw)
        r["confidence_score"] = 0.90 if (rut and bruto) else 0.60

        resultados.append(r)

    logger.info("PREVIRED parseado: %d trabajadores en %s", len(resultados), nombre_archivo)
    return resultados if resultados else [_resultado_vacio(nombre_archivo)]


# ─────────────────────────────────────────────────────────────────────
# CARTOLA BANCARIA en Excel
# ─────────────────────────────────────────────────────────────────────

def _parsear_cartola(wb: openpyxl.Workbook, nombre_archivo: str) -> list[dict]:
    """
    Parsea cartola bancaria en Excel.
    Columnas esperadas: Fecha | Descripción | Cargo | Abono | Saldo
    """
    resultados = []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = 0
    for i, row in enumerate(rows[:15]):
        row_str = " ".join(str(c).lower() for c in row if c)
        if ("fecha" in row_str and ("cargo" in row_str or "abono" in row_str)):
            header_idx = i
            break

    headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]]
    col_fecha = _find_col(headers, ["fecha"])
    col_desc  = _find_col(headers, ["descripción", "descripcion", "glosa", "detalle"])
    col_cargo = _find_col(headers, ["cargo", "débito", "debito", "monto"])
    col_abono = _find_col(headers, ["abono", "crédito", "credito"])
    col_saldo = _find_col(headers, ["saldo"])

    for row in rows[header_idx + 1:]:
        if not any(row):
            continue

        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        fecha = _parse_fecha_any(cell(col_fecha))
        if not fecha:
            continue

        cargo = _safe_int(cell(col_cargo))
        abono = _safe_int(cell(col_abono))
        saldo = _safe_int(cell(col_saldo))
        monto = -(cargo or 0) if cargo else (abono or 0)  # cargos = negativos

        r = _resultado_vacio(nombre_archivo)
        r["tipo"]             = TipoDocumento.CARTOLA
        r["rut_receptor"]     = _normalizar_rut(RUT_PARTIDO)
        r["fecha_documento"]  = fecha
        r["monto_bruto"]      = monto
        r["monto_liquido"]    = monto
        r["concepto"]         = str(cell(col_desc) or "Sin descripción")[:300]
        r["confidence_score"] = 0.85 if (fecha and monto) else 0.50
        resultados.append(r)

    logger.info("Cartola Excel parseada: %d movimientos en %s", len(resultados), nombre_archivo)
    return resultados if resultados else [_resultado_vacio(nombre_archivo)]


# ─────────────────────────────────────────────────────────────────────
# GENÉRICO — planilla de honorarios u otro Excel no clasificado
# ─────────────────────────────────────────────────────────────────────

def _parsear_generico(wb: openpyxl.Workbook, base: dict, nombre_archivo: str) -> list[dict]:
    """Parseo básico para Excel genérico — extrae lo que pueda encontrar."""
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    ruts_encontrados = []
    montos_encontrados = []

    for row in rows:
        for cell in row:
            val = str(cell or "").strip()
            if RE_RUT.match(val):
                ruts_encontrados.append(_normalizar_rut(val))
            m = RE_MONTO.match(val)
            if m:
                parsed = _safe_int(m.group(1).replace(".", "").replace(",", ""))
                if parsed and parsed > 1000:
                    montos_encontrados.append(parsed)

    base["rut_emisor"]     = ruts_encontrados[0] if ruts_encontrados else None
    base["monto_bruto"]    = max(montos_encontrados) if montos_encontrados else None
    base["confidence_score"] = 0.50 if base["rut_emisor"] else 0.30
    return [base]


# ─────────────────────────────────────────────────────────────────────
# Utilidades
# ─────────────────────────────────────────────────────────────────────

def _detectar_tipo(nombre: str, wb: openpyxl.Workbook) -> str:
    n = nombre.lower()
    if "previred" in n or "cotizaciones" in n:
        return TipoDocumento.PREVIRED
    if any(k in n for k in ["cartola", "extracto", "movimientos"]):
        return TipoDocumento.CARTOLA
    if "honorario" in n:
        return TipoDocumento.BHE
    # Revisar contenido de la primera hoja
    ws = wb.active
    for row in list(ws.iter_rows(values_only=True))[:5]:
        row_str = " ".join(str(c).lower() for c in row if c)
        if "previred" in row_str:
            return TipoDocumento.PREVIRED
        if "cartola" in row_str or "extracto" in row_str:
            return TipoDocumento.CARTOLA
    return TipoDocumento.OTRO


def _find_col(headers: list[str], keys: list[str]) -> Optional[int]:
    for key in keys:
        for i, h in enumerate(headers):
            if key in h:
                return i
    return None


def _resultado_vacio(nombre_archivo: str) -> dict:
    return {
        "tipo": TipoDocumento.OTRO,
        "rut_emisor": None,
        "nombre_emisor": None,
        "rut_receptor": _normalizar_rut(RUT_PARTIDO),
        "folio": None,
        "fecha_documento": None,
        "monto_bruto": None,
        "monto_neto": None,
        "monto_retencion": None,
        "monto_iva": None,
        "monto_liquido": None,
        "concepto": None,
        "confidence_score": 0.0,
        "archivo_nombre": nombre_archivo,
        "archivo_tipo": "xlsx",
    }


def _normalizar_rut(rut: str) -> str:
    rut = str(rut).strip().upper().replace(".", "").replace(" ", "")
    if "-" not in rut and len(rut) > 1:
        rut = rut[:-1] + "-" + rut[-1]
    return rut


def _safe_int(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val).replace(".", "").replace(",", "").replace("$", "").strip()))
    except (ValueError, TypeError):
        return None


def _parse_fecha_any(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    from dateutil import parser as dparser
    try:
        return dparser.parse(str(val)).date()
    except Exception:
        return None


def _parse_periodo(val: Any) -> Optional[date]:
    """Convierte "2026-05" o "Mayo 2026" a una fecha (día 1 del mes)."""
    if val is None:
        return None
    s = str(val).strip()
    try:
        from dateutil import parser as dparser
        return dparser.parse(s).date()
    except Exception:
        return None


def _parse_fecha(raw: str) -> Optional[date]:
    from dateutil import parser as dparser
    try:
        return dparser.parse(raw.strip()).date()
    except Exception:
        return None

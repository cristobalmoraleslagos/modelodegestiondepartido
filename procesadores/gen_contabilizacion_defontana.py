# -*- coding: utf-8 -*-
"""Genera material de contabilizacion para Defontana (anos sin contabilizar).
Usa M12 (gastos) + BHE (honorarios) mapeados a las cuentas REALES de Defontana
(extraidas de los balances 2024/2025). Salida: Excel con mapa + asientos borrador."""
import json, csv, glob, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ANIO = 2023   # ano a contabilizar
def num(s):
    s=(s or '').strip().replace('.','').replace(',','')
    try: return int(s)
    except: return 0
MESES=['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

# ── Mapa M12 (item SERVEL) -> cuenta Defontana real (codigo, nombre, contrapartida)
# Cuentas extraidas de los balances Defontana 2024/2025. 'VALIDAR' = revisar con contador.
MAPA = {
 'Gastos de Personal':                          ('4.5.1040.10.01','REMUNERACIONES','2.1.2030.20.01','REMUNERACIONES POR PAGAR'),
 'Adquisici':                                    ('4.9.1000.10.01','GASTOS A DISTRIBUIR','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Otros gastos de Administraci':                 ('4.1.2010.10.05','OTROS GASTOS DE ADMINISTRACION','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Fomento a Participaci':                        ('4.1.1010.10.08','GASTOS FOMENTO PARTICIPACION FEMENINA','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Actividades 10% mujer':                        ('4.1.1010.10.08','GASTOS FOMENTO PARTICIPACION FEMENINA','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Investigaci':                                  ('4.9.1000.10.01','GASTOS A DISTRIBUIR (VALIDAR: investigacion)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Educaci':                                      ('4.9.1000.10.01','GASTOS A DISTRIBUIR (VALIDAR: educacion civica)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Juvenil':                                      ('4.9.1000.10.01','GASTOS A DISTRIBUIR (VALIDAR: juvenil)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Preparaci':                                    ('4.1.3010.10.03','PREPARACION DE CANDIDATOS (electoral - VALIDAR cargo)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Publicidad Electoral':                         ('4.1.3010.10.04','GASTOS ELECTORAL (VALIDAR cargo)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Concejales':                       ('4.1.3010.10.05','GASTOS ELECTORAL CONCEJAL','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Consejeros Regionales':            ('4.1.3010.10.09','GASTO ELECTORAL CORE','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Alcaldes':                         ('4.1.3010.10.04','GASTOS ELECTORAL ALCALDE','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Gobernadores Regionales':          ('4.1.3010.10.10','GASTO ELECTORAL GOBERNADORES REGIONALES','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Presidencial':                     ('4.1.3010.10.03','GASTOS PRIMARIAS/ELECTORAL PRESIDENTE (VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Primarias Presidencial':           ('4.1.3010.10.12','GASTOS PRIMARIA PRESIDENCIAL','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Campaña Diputados':                        ('4.1.3010.10.14','GASTOS DIPUTADO','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gasto Electoral Diputados':                    ('4.1.3010.10.14','GASTOS DIPUTADO','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gasto Electoral Senadores':                    ('4.1.3010.10.13','GASTOS SENADOR','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gasto Electoral Presidencial':                 ('4.1.3010.10.03','GASTO ELECTORAL PRESIDENTE (VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Honorarios Campa':                             ('4.5.1030.10.01','HONORARIOS (campana)','2.1.2030.30.01','HONORARIOS POR PAGAR'),
 'Material Grafico Campa':                       ('4.1.3010.10.14','GASTOS ELECTORAL (VALIDAR cargo)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gastos Menores Campa':                         ('4.1.3010.10.14','GASTOS ELECTORAL (VALIDAR cargo)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Otros gastos Electorales':                     ('4.1.3010.10.14','GASTOS ELECTORAL (VALIDAR cargo)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Aporte Electoral':                             ('4.1.3010.10.14','APORTE A CANDIDATO (VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Aporte Campa':                                 ('4.1.3010.10.14','APORTE A CANDIDATO (VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Aporte Candidato':                             ('4.1.3010.10.14','APORTE A CANDIDATO (VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Eventos Partidarios':                          ('4.5.1030.10.30','EVENTOS PARTIDARIOS','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gastos Notariales':                            ('4.5.1030.10.19','GTOS. NOTARIALES','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Gastos por Prestamos':                         ('4.5.1020.10.03','Otros Intereses (gasto financiero)','2.1.1010.10.02','Credito (banco)'),
 # ── NO son gasto (van a balance, no a resultado): se marcan aparte ──
 'Transferencias entre cuentas':                 ('NO-RESULTADO','Movimiento entre bancos (no es gasto)','',''),
 'Cheque en Garant':                             ('NO-RESULTADO','Cheque en garantia (ACTIVO/garantia - no es gasto)','',''),
 'Cheque Devuelto':                              ('NO-RESULTADO','Activo/movimiento (no es gasto)','',''),
 'Devoluci':                                     ('NO-RESULTADO','Movimiento bancario (no es gasto)','',''),
 'Reintegros':                                   ('NO-RESULTADO','Devolucion de gasto (activo)','',''),
 'Anticipo Proveedores':                         ('1.1.1100.20.01','Anticipo a proveedores (ACTIVO)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Activo fijo':                                  ('1.2.1210.30.03','MUEBLES Y UTILES (ACTIVO)','2.1.1070.20.01','PROVEEDORES NACIONALES'),
 'Provisión de finiquito':                   ('4.5.1040.10.03','Provision (gasto) / Finiquitos por pagar','2.1.2030.20.02','FINIQUITOS POR PAGAR'),
 'Creditos':                                     ('NO-RESULTADO','Pasivo: credito recibido (no es gasto)','',''),
 'Créditos':                                 ('NO-RESULTADO','Pasivo: credito recibido (no es gasto)','',''),
}
def mapear(item):
    for k,v in MAPA.items():
        if k.lower() in item.lower():
            return v
    return ('4.9.1000.10.01','GASTOS A DISTRIBUIR (SIN MAPEO - VALIDAR)','2.1.1070.20.01','PROVEEDORES NACIONALES')

# ── Cargar M12 del ano ──
f = os.path.join(ROOT,'procesadores','output','M12_Gastos',f'{ANIO}-4.csv')
items = {}
with open(f, encoding='utf-8-sig') as fh:
    for row in csv.DictReader(fh, delimiter=';'):
        it=(row.get('Item de Gastos') or '').strip()
        por_mes={MESES[i]: num(row.get(MESES[i])) for i in range(12)}
        if sum(por_mes.values())>0: items[it]=por_mes

# ── Cargar BHE del ano (honorarios reales por mes) ──
bhe = json.load(open(os.path.join(ROOT,'procesadores','output','bhe_todas.json'),encoding='utf-8'))
hon_mes=defaultdict(int)
for b in bhe:
    if b['anio']==ANIO and b.get('estado','').upper()=='VIGENTE':
        hon_mes[b['mes']]+=b['honorario_bruto']

# ── Construir Excel ──
wb=Workbook()
HDR=PatternFill('solid',fgColor='9B2335'); WH=Font(color='FFFFFF',bold=True); B=Font(bold=True)
def hoja(ws, headers, rows, widths):
    for i,h in enumerate(headers,1):
        c=ws.cell(1,i,h); c.fill=HDR; c.font=WH; c.alignment=Alignment(horizontal='center')
    for r in rows:
        ws.append(r)
    for i,w in enumerate(widths,1):
        ws.column_dimensions[chr(64+i)].width=w

# Hoja 1: Mapa de cuentas
ws1=wb.active; ws1.title='Mapa de Cuentas'
rows1=[]
for it in sorted(items):
    cta,nom,ctra,ctra_nom = mapear(it)
    total=sum(items[it].values())
    rows1.append([it, total, cta, nom, ctra, ctra_nom])
hoja(ws1,['Item M12 SERVEL','Total '+str(ANIO),'Cuenta Defontana (Debe)','Nombre cuenta','Contrapartida (Haber)','Nombre contrapartida'],rows1,[42,16,18,38,18,28])

# Hoja 2: Asientos borrador (uno por item-mes, gasto vs contrapartida)
ws2=wb.create_sheet('Asientos 2023 (borrador)')
rows2=[]; nfolio=0
for it in sorted(items):
    cta,nom,ctra,ctra_nom = mapear(it)
    if cta=='NO-RESULTADO':
        for mes,v in items[it].items():
            if v>0:
                rows2.append([it,mes,'REVISAR','('+nom+')',v,v,'Movimiento NO de resultado - clasificar con contador'])
        continue
    for mes,v in items[it].items():
        if v>0:
            nfolio+=1
            rows2.append([nfolio,f'{mes} {ANIO}',cta,nom,v,0,'Gasto '+it[:30]])
            rows2.append([nfolio,f'{mes} {ANIO}',ctra,ctra_nom,0,v,'Contrapartida'])
hoja(ws2,['Folio','Periodo','Cuenta','Descripcion','Debe','Haber','Glosa'],rows2,[8,14,16,40,16,16,40])

# Hoja 3: Honorarios desde BHE (mas preciso que M12)
ws3=wb.create_sheet('Honorarios BHE 2023')
rows3=[]
for m in range(1,13):
    if hon_mes[m]>0:
        rows3.append([f'{MESES[m-1]} {ANIO}','4.5.1030.10.01','HONORARIOS',hon_mes[m],0,'Honorarios mes (BHE SII)'])
        rows3.append([f'{MESES[m-1]} {ANIO}','2.1.2030.30.01','HONORARIOS POR PAGAR',0,hon_mes[m],'Contrapartida'])
hoja(ws3,['Periodo','Cuenta','Descripcion','Debe','Haber','Glosa'],rows3,[14,16,28,16,16,30])

# Hoja 3b: Ya contabilizado en Defontana 2023 (NO recontabilizar)
ws3b=wb.create_sheet('Ya en Defontana 2023')
ya_rows=[
 ['30/12/2023','231200001','Cpra_FCA','FCA#313 CASA HOGAR CHARITO SPA (76.807.683-9)','1.1.1090.10.01','IVA CREDITO FISCAL',213750,0],
 ['30/12/2023','231200001','Cpra_FCA','FCA#313 CASA HOGAR CHARITO SPA (76.807.683-9)','4.1.1010.10.05','OTROS GASTOS DE ADMINISTRACION',1125000,0],
 ['30/12/2023','231200001','Cpra_FCA','FCA#313 CASA HOGAR CHARITO SPA (76.807.683-9)','2.1.1070.20.01','PROVEEDORES NACIONALES',0,1338750],
 ['29/04/2023','22','EGRESO','ABONO DEUDA N.XX SIGLO XXI','1.1.1100.20.01','Anticipo a proveedores',400000,0],
 ['29/04/2023','22','EGRESO','ABONO DEUDA N.XX SIGLO XXI','1.1.1010.20.04','BANCO BCI 13950223',0,400000],
 ['29/04/2023','27','EGRESO','ABONO PROVEEDORES N.XX SIGLO XXI','1.1.1100.20.01','Anticipo a proveedores',450000,0],
 ['29/04/2023','27','EGRESO','ABONO PROVEEDORES N.XX SIGLO XXI','1.1.1010.20.04','BANCO BCI 13950223',0,450000],
]
hoja(ws3b,['Fecha','Comprobante','Tipo','Glosa','Cuenta','Nombre cuenta','Debe','Haber'],ya_rows,[12,12,10,42,16,30,14,14])
ws3b.append([])
ws3b.append(['NOTA: estos 3 comprobantes (~$2.19M) YA estan en Defontana 2023. NO recontabilizar.'])

# Hoja 3c: Control de totales (cuadratura)
ws3c=wb.create_sheet('Control de Totales')
tot_m12=sum(sum(v.values()) for v in items.values())
tot_hon=sum(hon_mes.values())
ctrl=[
 ['Concepto','Monto $','Fuente'],
 ['Gastos M12 SERVEL 2023 (todos los items)',tot_m12,'M12 output 2023-4.csv'],
 ['  de los cuales NO-RESULTADO (no van a gasto)',sum(sum(items[it].values()) for it in items if mapear(it)[0]=='NO-RESULTADO'),'clasificar a balance'],
 ['Honorarios BHE 2023 (document-level, VIGENTES)',tot_hon,'SII BHE bhe_todas.json'],
 ['Boletas BHE 2023 vigentes (cantidad)',len([b for b in bhe if b['anio']==ANIO and b.get('estado','').upper()=='VIGENTE']),'SII'],
 ['Ya contabilizado en Defontana 2023',2188750,'Libro Mayor Defontana 2023'],
 ['','',''],
 ['ADVERTENCIA: M12 "Gastos de Personal" incluye remuneraciones Y honorarios.','',''],
 ['No sumar BHE encima sin que el contador separe la parte de honorarios.','',''],
]
hoja(ws3c,['Concepto','Monto $','Fuente'],ctrl[1:],[55,18,30])

# Hoja 3d: Ingresos M6 2023 (plantilla con cuentas REALES de Defontana)
# Cuentas de ingreso extraidas del Libro Mayor Defontana 2024/2025 (grupo 3).
# 2023 NO tiene ingresos contabilizados; el monto debe obtenerse de las CARTOLAS
# bancarias / registros de Tesoreria / formulario M6 que el partido haya presentado.
ws3d=wb.create_sheet('Ingresos M6 2023 (plantilla)')
ING_CUENTAS=[
 ('3.1.1010.10.01','INGRESOS PROCEDENTES DE APORTES TRIMESTRAL (aporte estatal SERVEL)','$345.788.634 (Transparencia m11; monto exacto a reconfirmar con cartola)'),
 ('3.1.2010.10.01','INGRESOS DE COTIZACIONES ORDINARIAS AFILIADOS','de cartola/Tesoreria'),
 ('3.1.2010.10.02','INGRESOS DE COTIZACIONES EXTRAORDINARIAS AFILIADOS','de cartola/Tesoreria'),
 ('3.1.2010.10.03','INGRESOS DE COTIZACIONES NO AFILIADOS / DONACIONES','de cartola/Tesoreria'),
 ('3.1.3010.10.01','INGRESOS PROVENIENTES DEL PROPIO PATRIMONIO','de cartola/Tesoreria'),
 ('3.1.4010.xx.xx','INGRESOS FINANCIAMIENTO/REEMBOLSO ELECTORAL (por cargo)','2023: Consejo Constitucional - de SERVEL'),
 ('3.5.0100.10.01','INTERESES EFECTIVAMENTE PERCIBIDOS','de cartola bancaria'),
]
rows3d=[[c,n,0,'',obs] for c,n,obs in ING_CUENTAS]
hoja(ws3d,['Cuenta Defontana','Nombre cuenta','Monto $ (llenar)','Fecha','Fuente / Observacion'],rows3d,[18,52,16,12,42])
ws3d.append([])
ws3d.append(['BRECHA DE FINANCIAMIENTO 2023:'])
ws3d.append([f'Gastos M12 2023 = {tot_m12:,}  |  Ingresos contabilizados 2023 = 0  |  Aporte estatal = 0'])
ws3d.append(['=> Todo el gasto 2023 se financio con cotizaciones / reembolsos electorales /'])
ws3d.append(['   prestamos / patrimonio. ESTE es el dato que falta y debe salir de las cartolas.'])

# Hoja 4: Instrucciones
ws4=wb.create_sheet('LEER PRIMERO')
notas=[
 ['MATERIAL DE CONTABILIZACION '+str(ANIO)+' - PARTIDO COMUNISTA DE CHILE'],
 [''],
 ['Origen: M12 Gastos SERVEL + BHE honorarios (SII), mapeados a las cuentas REALES'],
 ['de Defontana extraidas de los balances 2024 y 2025 ya contabilizados.'],
 [''],
 ['IMPORTANTE - este es un BORRADOR para validar con el contador:'],
 ['1. Las cuentas marcadas (VALIDAR) requieren confirmacion del contador.'],
 ['2. El M12 "Adquisicion de Bienes" es un cajon: el contador lo reparte a'],
 ['   arriendos, servicios basicos, comunicaciones, etc. (lo ideal: contabilizar'],
 ['   desde los documentos del RCV cargados en Compras, no desde el M12 agregado).'],
 ['3. Honorarios: usar la Hoja "Honorarios BHE 2023" (document-level, mas preciso).'],
 ['4. Los items "NO-RESULTADO" (transferencias, cheques, creditos) NO son gasto:'],
 ['   van al balance (bancos/pasivos), el contador los clasifica.'],
 ['5. La contrapartida sugerida es Proveedores Nacionales / Por Pagar; el pago'],
 ['   real (Haber Banco) es un asiento aparte por cartola.'],
 [''],
 ['RECOMENDADO: cargar primero los documentos 2023 del RCV en Compras de Defontana,'],
 ['y contabilizarlos con el plan de cuentas existente (que ya esta completo).'],
]
for r in notas: ws4.append(r)
ws4.column_dimensions['A'].width=95
ws4['A1'].font=Font(bold=True,size=13,color='9B2335')

out=os.path.join(ROOT,'DEFONTANA','Contabilizacion',f'Contabilizacion_{ANIO}_Defontana.xlsx')
os.makedirs(os.path.dirname(out),exist_ok=True)
wb.save(out)
print('OK ->',out)
print(f'  {len(items)} items M12 mapeados | {nfolio} lineas de asiento | honorarios {sum(hon_mes.values()):,}')
